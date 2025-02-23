# import streamlit as st
# import pandas as pd
# import json
# from openpyxl import load_workbook
# from openpyxl.styles import Font, Alignment

# # Function to process the data
# def process_data(file, selected_columns, names_json_path):
#     # Load the Excel file into a DataFrame
#     dataframe = pd.read_excel(file, sheet_name='Sheet1')

#     # Drop unselected columns
#     dataframe = dataframe[selected_columns]
#     dataframe.dropna(inplace=True, subset=['MARK'])

#     # Generate the 'Block' column for grouping
#     dataframe['Block'] = (dataframe['PCS/CTN'] != dataframe['PCS/CTN'].shift()).cumsum()

#     # Group data and generate summaries
#     grouped = dataframe.groupby(['MARK', 'Block']).agg(
#         First_CTN=('CTN NO', 'first'),
#         Last_CTN=('CTN NO', 'last'),
#         T_CTN=('CTN/TOTAL', 'sum'),
#         WT=('WEIGHT/TOTAL', 'first'),
#         UNIT=('UNITS', 'first'),
#         QTY=('PCS/CTN', 'first'),
#         T_QTY=('PCS/CTN', 'sum'),
#         DESCRIPTION=('DESCRIPTION', 'first')
#     ).reset_index()

#     # Generate 'CTN NO' field
#     grouped['CTN NO'] = grouped.apply(
#         lambda row: (
#             f"{row['First_CTN'] if row['First_CTN'] is not None else 1} - "
#             f"{row['Last_CTN'] if row['Last_CTN'] is not None else int(row['T_CTN'])}"
#             if row['T_CTN'] > 1
#             else f"{row['First_CTN'] if row['First_CTN'] is not None else 1}"
#         ),
#         axis=1
#     )

#     grouped['T.QTY'] = grouped['T_QTY']
#     grouped['T.CTN'] = grouped['T_CTN']

#     # Final columns to display
#     final_columns = ['MARK', 'CTN NO', 'DESCRIPTION', 'T.CTN', 'QTY', 'UNIT', 'T.QTY', 'WT']
#     final_dataset = grouped[final_columns]

#     # Load the names.json file
#     with open(names_json_path, 'r') as file:
#         name_mapping = json.load(file)

#     # Create a mapping dictionary
#     description_mapping = {item['rough']: item['formatted'] for item in name_mapping}

#     # Replace descriptions using the mapping
#     final_dataset['DESCRIPTION'] = final_dataset['DESCRIPTION'].replace(description_mapping)

#     # Check for missing names
#     unique_descriptions = final_dataset['DESCRIPTION'].unique()
#     missing_descriptions = [desc for desc in unique_descriptions if desc not in description_mapping.keys()]

#     return final_dataset, missing_descriptions

# # Function to update names.json
# def update_names_json(names_json_path, new_mapping):
#     with open(names_json_path, 'r') as file:
#         name_mapping = json.load(file)
#     name_mapping.extend(new_mapping)
#     with open(names_json_path, 'w') as file:
#         json.dump(name_mapping, file, indent=4)

# # Streamlit UI
# st.title("Excel File Processor")

# # File upload
# uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])
# if uploaded_file is not None:
#     # Load and display the columns
#     dataframe = pd.read_excel(uploaded_file, sheet_name='Sheet1')
#     st.write("Select the required columns:")
#     columns = dataframe.columns.tolist()
#     selected_columns = st.multiselect("Columns", columns, default=columns)

#     if st.button("Process Data"):
#         if not selected_columns:
#             st.error("Please select at least one column.")
#         else:
#             # Process data
#             names_json_path = "names.json"
#             processed_data, missing_names = process_data(uploaded_file, selected_columns, names_json_path)

#             # Display processed data
#             st.write("Processed Data Preview:")
#             st.dataframe(processed_data)

#             # Handle missing names
#             if missing_names:
#                 st.warning("Some names are missing in the dataset:")
#                 for name in missing_names:
#                     formatted_name = st.text_input(f"Formatted name for '{name}'", key=name)
#                     if formatted_name:
#                         new_mapping = [{"rough": name, "formatted": formatted_name}]
#                         update_names_json(names_json_path, new_mapping)
#                         st.success(f"Name '{name}' updated to '{formatted_name}' in names.json.")

#             # Save the processed data
#             processed_file_path = "processed.xlsx"
#             processed_data.to_excel(processed_file_path, index=False)

#             st.success(f"Data processed and saved to {processed_file_path}.")


# import pandas as pd
# import json
# import streamlit as st
# from openpyxl import load_workbook
# from openpyxl.styles import Font, Alignment

# # Load names.json
# def load_name_mapping(file_path):
#     with open(file_path, 'r') as file:
#         return json.load(file)

# def save_name_mapping(file_path, name_mapping):
#     with open(file_path, 'w') as file:
#         json.dump(name_mapping, file, indent=4)

# # Function to process the dataset
# def process_dataset(dataframe, name_mapping):
#     description_mapping = {item['rough']: item['formatted'] for item in name_mapping}
#     dataframe['DESCRIPTION'] = dataframe['DESCRIPTION'].replace(description_mapping)
#     dataframe['Block'] = (dataframe['PCS/CTN'] != dataframe['PCS/CTN'].shift()).cumsum()
#     grouped = dataframe.groupby(['MARK', 'Block']).agg(
#         First_CTN=('CTN NO', 'first'),
#         Last_CTN=('CTN NO', 'last'),
#         T_CTN=('CTN/TOTAL', 'sum'),
#         WT=('WEIGHT/TOTAL', 'first'),
#         UNIT=('UNITS', 'first'),
#         QTY=('PCS/CTN', 'first'),
#         T_QTY=('PCS/CTN', 'sum'),
#         DESCRIPTION=('DESCRIPTION', 'first')
#     ).reset_index()

#     grouped['CTN NO'] = grouped.apply(
#         lambda row: (
#             f"{row['First_CTN'] if row['First_CTN'] is not None else 1} - "
#             f"{row['Last_CTN'] if row['Last_CTN'] is not None else int(row['T_CTN'])}"
#             if row['T_CTN'] > 1
#             else f"{row['First_CTN'] if row['First_CTN'] is not None else 1}"
#         ),
#         axis=1
#     )
#     grouped['T.QTY'] = grouped['T_QTY']
#     grouped['T.CTN'] = grouped['T_CTN']
#     final_columns = ['MARK', 'CTN NO', 'DESCRIPTION', 'T.CTN', 'QTY', 'UNIT', 'T.QTY', 'WT']
#     final_dataset = grouped[final_columns]
#     return final_dataset

# # Main Streamlit app
# st.title("Excel Processing App")

# # Step 1: Upload Excel file
# uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])

# if uploaded_file:
#     dataframe = pd.read_excel(uploaded_file)
#     st.write("Preview of uploaded data:")
#     st.dataframe(dataframe.head())

#     # Step 2: Column selection
#     selected_columns = st.multiselect("Select the required columns", dataframe.columns.tolist())
    
#     if selected_columns:
#         dataframe = dataframe[selected_columns]
#         st.write("Preview of selected columns:")
#         st.dataframe(dataframe.head())
    
#         # Step 3: Process the dataset
#         name_mapping = load_name_mapping("names.json")
#         processed_dataset = process_dataset(dataframe, name_mapping)
        
#         # Step 4: Check for missing names
#         missing_names = processed_dataset[~processed_dataset['DESCRIPTION'].isin(
#             [item['formatted'] for item in name_mapping]
#         )]['DESCRIPTION'].unique()
        
#         if missing_names.size > 0:
#             st.warning("Some descriptions are not mapped:")
#             for name in missing_names:
#                 new_name = st.text_input(f"Enter formatted name for '{name}':")
#                 if new_name:
#                     # Append the new name to the name mapping
#                     name_mapping.append({"rough": name, "formatted": new_name})
#                     save_name_mapping("names.json", name_mapping)
#                     st.success(f"'{name}' mapped to '{new_name}' and saved.")

#         # Step 5: Display processed data
#         st.write("Preview of processed data:")
#         st.dataframe(processed_dataset)

#         # Step 6: Save processed data
#         save_path = st.text_input("Enter a filename to save the processed file (e.g., processed.xlsx):")
#         if st.button("Save Processed Data"):
#             processed_dataset.to_excel(save_path, index=False)
#             st.success(f"Processed file saved as {save_path}!")

import os
import pandas as pd
import json
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# Load names.json
def load_name_mapping(file_path):
    # Check if the file exists
    if not os.path.exists(file_path):
        # If the file doesn't exist, create it with an empty list
        with open(file_path, 'w') as file:
            json.dump([], file, indent=4)
    # Load the JSON file
    with open(file_path, 'r') as file:
        return json.load(file)

# Save the name mapping back to the file
def save_name_mapping(file_path, name_mapping):
    with open(file_path, 'w') as file:
        json.dump(name_mapping, file, indent=4)

# Function to process the dataset
def process_dataset(dataframe, name_mapping):
    dataframe.dropna();
    description_mapping = {item['rough']: item['formatted'] for item in name_mapping}
    dataframe['DESCRIPTION'] = dataframe['DESCRIPTION'].replace(description_mapping)
    dataframe['Block'] = (dataframe['PCS/CTN'] != dataframe['PCS/CTN'].shift()).cumsum()
    grouped = dataframe.groupby(['MARK', 'Block']).agg(
        First_CTN=('CTN NO', 'first'),
        Last_CTN=('CTN NO', 'last'),
        T_CTN=('CTN/TOTAL', 'sum'),
        WT=('WEIGHT/TOTAL', 'first'),
        UNIT=('UNITS', 'first'),
        QTY=('PCS/CTN', 'first'),
        T_QTY=('PCS/CTN', 'sum'),
        DESCRIPTION=('DESCRIPTION', 'first')
    ).reset_index()

    grouped['CTN NO'] = grouped.apply(
        lambda row: (
            f"{row['First_CTN'] if row['First_CTN'] is not None else 1} - "
            f"{row['Last_CTN'] if row['Last_CTN'] is not None else int(row['T_CTN'])}"
            if row['T_CTN'] > 1
            else f"{row['First_CTN'] if row['First_CTN'] is not None else 1}"
        ),
        axis=1
    )
    grouped['T.QTY'] = grouped['T_QTY']
    grouped['T.CTN'] = grouped['T_CTN']
    final_columns = ['MARK', 'CTN NO', 'DESCRIPTION', 'T.CTN', 'QTY', 'UNIT', 'T.QTY', 'WT']
    final_dataset = grouped[final_columns]
    return final_dataset

# Main Streamlit app
st.title("Excel Processing App")

# Step 1: Upload Excel file
uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])

if uploaded_file:
    dataframe = pd.read_excel(uploaded_file)
    st.write("Preview of uploaded data:")
    st.dataframe(dataframe.head())

    # Step 2: Column selection
    selected_columns = st.multiselect("Select the required columns", dataframe.columns.tolist())
    
    if selected_columns:
        dataframe = dataframe[selected_columns]
        st.write("Preview of selected columns:")
        st.dataframe(dataframe.head())
    
        # Step 3: Process the dataset
        name_mapping = load_name_mapping("names.json")
        processed_dataset = process_dataset(dataframe, name_mapping)
        
        # Step 4: Check for missing names
        missing_names = processed_dataset[~processed_dataset['DESCRIPTION'].isin(
            [item['formatted'] for item in name_mapping]
        )]['DESCRIPTION'].unique()
        
        if missing_names.size > 0:
            st.warning("Some descriptions are not mapped:")
            for name in missing_names:
                new_name = st.text_input(f"Enter formatted name for '{name}':")
                if new_name:
                    # Append the new name to the name mapping
                    name_mapping.append({"rough": name, "formatted": new_name})
                    save_name_mapping("names.json", name_mapping)
                    st.success(f"'{name}' mapped to '{new_name}' and saved.")

        # Step 5: Display processed data
        st.write("Preview of processed data:")
        st.dataframe(processed_dataset)

        # Step 6: Save processed data
        save_path = st.text_input("Enter a filename to save the processed file (e.g., processed.xlsx):")
        if st.button("Save Processed Data"):
            # processed_dataset.to_excel(save_path, index=False)
            st.success(f"Processed file saved as {save_path}!")

            consolidated = processed_dataset.groupby(['DESCRIPTION'])
            consolidated_dataset = consolidated.agg(
                Description=('DESCRIPTION', 'first'),
                Quantity=('T.QTY', 'sum'),
            )

            # Write the consolidated data back to the Excel file
            # Append the consolidated dataset to the Excel file as before
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Alignment

            # Load the existing Excel file
            file_path = 'processed.xlsx'  # Replace with your file path
            workbook = load_workbook(file_path)

            # Load the first sheet (or specify the sheet name)
            sheet_name = workbook.sheetnames[0]
            sheet = workbook[sheet_name]

            cambria_font = Font(name='Cambria', size=11)
            for row in sheet.iter_rows():
                for cell in row:
                    # Set the font
                    cell.font = cambria_font
                    cell.border = None
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Define the starting cell for consolidated data
            start_row = 6
            start_col = 10  # Column 'J' corresponds to the 10th column
            data = consolidated_dataset.values.tolist()
            print(data)
            headers = consolidated_dataset.columns.tolist()

            # Write data
            for row_idx, row in enumerate(data, start=start_row):
                for col_idx, value in enumerate(row, start=start_col):
                    sheet.cell(row=row_idx, column=col_idx, value=value)

            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter  # Get column letter (e.g., A, B, C)
                for cell in column:
                    if cell.value:  # Check if cell is not empty
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = max_length + 2  # Add extra space for padding
                sheet.column_dimensions[column_letter].width = adjusted_width

            # Save the workbook
            workbook.save(save_path)

            print("Data updated and saved successfully!")

# import os
# import pandas as pd
# import json
# import streamlit as st
# from openpyxl import load_workbook
# from openpyxl.styles import Font, Alignment

# # Load names.json
# def load_name_mapping(file_path):
#     # Check if the file exists
#     if not os.path.exists(file_path):
#         # If the file doesn't exist, create it with an empty list
#         with open(file_path, 'w') as file:
#             json.dump([], file, indent=4)
#     # Load the JSON file
#     with open(file_path, 'r') as file:
#         return json.load(file)

# # Save the name mapping back to the file
# def save_name_mapping(file_path, name_mapping):
#     with open(file_path, 'w') as file:
#         json.dump(name_mapping, file, indent=4)

# # Function to process the dataset
# def process_dataset(dataframe, name_mapping):
#     # dataframe = dataframe.dropna()
#     description_mapping = {item['rough']: item['formatted'] for item in name_mapping}
#     dataframe['DESCRIPTION'] = dataframe['DESCRIPTION'].replace(description_mapping)
#     dataframe['Block'] = (dataframe['PCS/CTN'] != dataframe['PCS/CTN'].shift()).cumsum()
#     grouped = dataframe.groupby(['MARK', 'Block']).agg(
#         First_CTN=('CTN NO', 'first'),
#         Last_CTN=('CTN NO', 'last'),
#         T_CTN=('CTN/TOTAL', 'sum'),
#         WT=('WEIGHT/TOTAL', 'first'),
#         UNIT=('UNITS', 'first'),
#         QTY=('PCS/CTN', 'first'),
#         T_QTY=('PCS/CTN', 'sum'),
#         DESCRIPTION=('DESCRIPTION', 'first')
#     ).reset_index()

#     grouped['CTN NO'] = grouped.apply(
#         lambda row: (
#             f"{row['First_CTN'] if row['First_CTN'] is not None else 1} - "
#             f"{row['Last_CTN'] if row['Last_CTN'] is not None else int(row['T_CTN'])}"
#             if row['T_CTN'] > 1
#             else f"{row['First_CTN'] if row['First_CTN'] is not None else 1}"
#         ),
#         axis=1
#     )
#     grouped['T.QTY'] = grouped['T_QTY']
#     grouped['T.CTN'] = grouped['T_CTN']
#     final_columns = ['MARK', 'CTN NO', 'DESCRIPTION', 'T.CTN', 'QTY', 'UNIT', 'T.QTY', 'WT']
#     final_dataset = grouped[final_columns]
#     return final_dataset

# # Main Streamlit app
# st.title("Excel Processing App")

# # Step 1: Upload Excel file
# uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])

# if uploaded_file:
#     dataframe = pd.read_excel(uploaded_file)
#     st.write("Preview of uploaded data:")
#     st.dataframe(dataframe.head())

#     # Step 2: Display current column names
#     st.write("Column names in your uploaded file:")
#     current_columns = dataframe.columns.tolist()
#     st.write(current_columns)

#     # Step 3: Allow user to map columns
#     expected_columns = ["MARK", "DESCRIPTION", "CTN NO", "PCS/CTN", "CTN/TOTAL", "WEIGHT/TOTAL", "UNITS"]
#     column_mapping = {}
    
#     for expected_col in expected_columns:
#         column_mapping[expected_col] = st.selectbox(
#             f"Map column for '{expected_col}':",
#             options=["None"] + current_columns,
#             index=current_columns.index(expected_col) if expected_col in current_columns else 0,
#         )

#     # Apply column mapping
#     if st.button("Apply Column Mapping"):
#         dataframe = dataframe.rename(columns={v: k for k, v in column_mapping.items() if v != "None"})
#         missing_columns = [col for col in expected_columns if col not in dataframe.columns]
        
#         if missing_columns:
#             st.error(f"The following required columns are missing after mapping: {', '.join(missing_columns)}")
#         else:
#             st.success("Column mapping applied successfully!")
#             st.write("Updated dataframe with mapped columns:")
#             st.dataframe(dataframe.head())

#             # Step 4: Process the dataset
#             name_mapping = load_name_mapping("names.json")
#             processed_dataset = process_dataset(dataframe, name_mapping)

#             # Step 5: Check for missing names
#             missing_names = processed_dataset[~processed_dataset['DESCRIPTION'].isin(
#                 [item['formatted'] for item in name_mapping]
#             )]['DESCRIPTION'].unique()

#             if missing_names.size > 0:
#                 st.warning("Some descriptions are not mapped:")
#                 for name in missing_names:
#                     new_name = st.text_input(f"Enter formatted name for '{name}':")
#                     if new_name:
#                         # Append the new name to the name mapping
#                         name_mapping.append({"rough": name, "formatted": new_name})
#                         save_name_mapping("names.json", name_mapping)
#                         st.success(f"'{name}' mapped to '{new_name}' and saved.")

#             # Step 6: Display processed data
#             st.write("Preview of processed data:")
#             st.dataframe(processed_dataset)

#             # Step 7: Save processed data
#             save_path = st.text_input("Enter a filename to save the processed file (e.g., processed.xlsx):")
#             if st.button("Save Processed Data"):
#                 processed_dataset.to_excel(save_path, index=False)
#                 st.success(f"Processed file saved as {save_path}!")
