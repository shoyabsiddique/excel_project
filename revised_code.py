import pandas as pd
import numpy as np
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

def process_data(
    input_file,
    sheet_name,
    output_file,
    column_config,
    json_mapping_file,
    consolidated_start_row=6,
    consolidated_start_col=10,
    drop_columns=None
):
    """
    Process the Excel data dynamically based on the user-specified column configuration.
    Instead of dropping rows with missing values, this function fills them with default values.
    
    Parameters:
    - input_file: Path to the input Excel file.
    - sheet_name: Sheet name in the Excel file to read.
    - output_file: Path where the processed Excel file will be saved.
    - column_config: Dictionary mapping roles to column names. For example:
        {
            'mark': 'MARK',
            'description': 'DESCRIPTION',
            'pcs_per_ctn': 'PCS/CTN',
            'ctn_no': 'CTN NO',
            'ctn_total': 'CTN/TOTAL',
            'weight_total': 'WEIGHT/TOTAL',
            'units': 'UNITS'
        }
    - json_mapping_file: Path to the JSON file containing description mappings.
    - consolidated_start_row: Row number where consolidated data starts (default 6).
    - consolidated_start_col: Column index where consolidated data starts (default 10 for 'J').
    - drop_columns: List of columns to drop from the input DataFrame.
    """
    
    # Step 1: Read the Excel file and drop unwanted columns
    df = pd.read_excel(input_file, sheet_name=sheet_name)
    if drop_columns:
        df = df.drop(columns=drop_columns)
    
    # Instead of dropping rows where the mark column is missing,
    # fill missing values with a default value (e.g., "Unknown")
    df[column_config['mark']] = df[column_config['mark']].fillna("Unknown")
    
    # You can also fill missing values for other key columns as needed:
    # df[column_config['description']] = df[column_config['description']].fillna("No Description")
    # ... and so on
    
    # Create a 'Block' identifier based on changes in the 'pcs_per_ctn' column
    df['Block'] = (df[column_config['pcs_per_ctn']] != df[column_config['pcs_per_ctn']].shift()).cumsum()
    
    # Step 2: Group by the 'mark' and 'Block', then aggregate
    grouped = df.groupby([column_config['mark'], 'Block']).agg(
        First_CTN=(column_config['ctn_no'], 'first'),
        Last_CTN=(column_config['ctn_no'], 'last'),
        T_CTN=(column_config['ctn_total'], 'sum'),
        WT=(column_config['weight_total'], 'first'),
        UNITS=(column_config['units'], 'first'),
        QTY=(column_config['pcs_per_ctn'], 'first'),
        T_QTY=(column_config['pcs_per_ctn'], 'sum'),
        DESCRIPTION=(column_config['description'], 'first')
    ).reset_index()
    
    # Generate the CTN NO field dynamically,
    # using pd.notnull() to check for missing values
    grouped['CTN NO'] = grouped.apply(
        lambda row: (
            f"{row['First_CTN'] if pd.notnull(row['First_CTN']) else 1} - "
            f"{row['Last_CTN'] if pd.notnull(row['Last_CTN']) else int(row['T_CTN'])}"
            if row['T_CTN'] > 1
            else f"{row['First_CTN'] if pd.notnull(row['First_CTN']) else 1}"
        ),
        axis=1
    )
    
    # Rename columns for the final dataset
    grouped['T.QTY'] = grouped['T_QTY']
    grouped['T.CTN'] = grouped['T_CTN']
    
    final_columns = [
        column_config['mark'], 'CTN NO', column_config['description'],
        'T.CTN', 'QTY', column_config['units'], 'T.QTY', 'WT'
    ]
    final_dataset = grouped[final_columns]
    # Step 3: Replace descriptions using a JSON mapping
    with open(json_mapping_file, 'r') as file:
        name_mapping = json.load(file)
    description_mapping = {item['rough']: item['formatted'] for item in name_mapping}
    final_dataset[column_config['description']] = final_dataset[column_config['description']].replace(description_mapping)
    final_dataset = final_dataset[final_dataset[column_config['description']] != "Noen"]
    
    # Save the final dataset to Excel
    final_dataset.to_excel(output_file, index=False)
    
    # Step 4: Create a consolidated dataset based on the description column
    consolidated = final_dataset.groupby(column_config['description']).agg(
        Description=(column_config['description'], 'first'),
        Quantity=('T.QTY', 'sum')
    )
    
    # Append the consolidated dataset to the Excel file using openpyxl
    workbook = load_workbook(output_file)
    sheet = workbook[workbook.sheetnames[0]]
    
    # Apply styling to the entire sheet
    cambria_font = Font(name='Cambria', size=11)
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = cambria_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write consolidated data starting at the specified cell
    data = consolidated.reset_index(drop=True).values.tolist()
    for row_idx, row in enumerate(data, start=consolidated_start_row):
        for col_idx, value in enumerate(row, start=consolidated_start_col):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(output_file)
    print("Data updated and saved successfully!")

# Example usage:

# Define the column configuration mapping
column_config = {
    'mark': 'MARK',
    'description': 'DESCRIPTION',
    'pcs_per_ctn': 'PCS/CTN',
    'ctn_no': 'CTN NO',
    'ctn_total': 'CTN/TOTAL',
    'weight_total': 'WEIGHT/TOTAL',
    'units': 'UNITS'
}

# Optionally, specify columns to drop (if they are not needed for the analysis)
drop_columns = ['BIS NO.', 'BIS MODEL NO.', 'MAH', 'MADE IN', 'LOGO']

process_data(
    input_file='test.xlsx',
    sheet_name='Sheet1',
    output_file='processed_revised.xlsx',
    column_config=column_config,
    json_mapping_file='names.json',
    consolidated_start_row=6,
    consolidated_start_col=10,
    drop_columns=None
)
