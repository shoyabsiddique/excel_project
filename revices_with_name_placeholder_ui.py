import sys
import json
import pandas as pd
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                            QHBoxLayout, QPushButton, QLabel, QFileDialog, 
                            QComboBox, QScrollArea, QMessageBox, QLineEdit,
                            QDialog, QGridLayout)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont, QPalette, QColor, QIcon

import pandas as pd
import numpy as np
import json
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

def process_data(
    df,
    sheet_name,
    output_file,
    column_config,
    description_mapping,
    consolidated_start_row=6,
    consolidated_start_col=10,
    drop_columns=None
):
    # Step 1: Read the Excel file and drop unwanted columns.
    if drop_columns:
        df = df.drop(columns=drop_columns)
    
    # Fill missing values for the mark column.
    df[column_config['mark']] = df[column_config['mark']].fillna("Unknown")
    df[column_config['description']] = df[column_config['description']].fillna("None")
    df = df[df[column_config['description']] != "None"]
    # For any extra columns needed for placeholder substitution (like MAH), fill missing values if desired.
    if 'mah' in column_config:
        df[column_config['mah']] = df[column_config['mah']].fillna("0")
    print(description_mapping)

    import re

    def substitute_placeholders(row, mapping_list, col_config):
        """
        Returns a tuple: (final_description, consolidated_description)
        - final_description: The result of performing placeholder substitution.
        - consolidated_description: If any placeholders were used, then return the original rough description,
        otherwise return the substituted string.
        """
        raw = row[col_config['description']]
        for mapping in mapping_list:
            if mapping['rough'] == raw:
                formatted = mapping['formatted']
                # Use regex to find all placeholders
                pattern = re.compile(r'%([^%]+)%')
                def replacer(match):
                    placeholder = match.group(1)
                    if placeholder in row:
                        return str(row[placeholder])
                    else:
                        return match.group(0)  # Return the original placeholder if column not found
                new_desc = pattern.sub(replacer, formatted)
                # For consolidation, use the raw description if any substitution was made
                if '%' in new_desc:
                    return new_desc, raw
                else:
                    return new_desc, new_desc
        # If no matching rough description is found, return the raw description for both
        return raw, raw
    
    # Apply substitution to create two new columns: final description and consolidated description.
    results = df.apply(lambda row: pd.Series(substitute_placeholders(row, description_mapping, column_config)), axis=1)
    df[column_config['description']] = results[0]
    df['consolidated_desc'] = results[1]
    
    # (Rest of the function remains the same)
    # ...
    # Create a 'Block' identifier based on changes in the 'pcs_per_ctn' column.
    df['Block'] = (df[column_config['pcs_per_ctn']] != df[column_config['pcs_per_ctn']].shift()).cumsum()
    
    # Step 2: Group by the 'mark' and 'Block', then aggregate.
    # Include both the final description and the consolidated description in the aggregation.
    grouped = df.groupby([column_config['mark'], 'Block']).agg(
        First_CTN=(column_config['ctn_no'], 'first'),
        Last_CTN=(column_config['ctn_no'], 'last'),
        T_CTN=(column_config['ctn_total'], 'sum'),
        WT=(column_config['weight_total'], 'first'),
        UNITS=(column_config['units'], 'first'),
        QTY=(column_config['pcs_per_ctn'], 'first'),
        T_QTY=(column_config['pcs_per_ctn'], 'sum'),
        DESCRIPTION=(column_config['description'], 'first'),
        CONSOLIDATED_DESC=('consolidated_desc', 'first')
    ).reset_index()
    
    # Generate the CTN NO field dynamically.
    # grouped['CTN NO'] = grouped.apply(
    #     lambda row: (
    #         f"{row['First_CTN'] if pd.notnull(row['First_CTN']) else 1} - "
    #         f"{row['Last_CTN'] if pd.notnull(row['Last_CTN']) else int(row['T_CTN'])}"
    #         if row['T_CTN'] > 1
    #         else f"{row['First_CTN'] if pd.notnull(row['First_CTN']) else 1}"
    #     ),
    #     axis=1
    # )
    
    grouped['CTN NO'] = grouped.apply(
    lambda row: (
        f"{row['First_CTN'] if pd.notnull(row['First_CTN']) else 1} - "
        f"{''.join(filter(str.isdigit, str(row['Last_CTN']))) if pd.notnull(row['Last_CTN']) else int(row['T_CTN'])}"
        if row['T_CTN'] > 1
        else f"{row['First_CTN'] if pd.notnull(row['First_CTN']) else 1}"
    ),
    axis=1
)
    
    # Rename columns for the final dataset.
    grouped['T.CTN'] = grouped['T_CTN']
    grouped['T.QTY'] = grouped['T.CTN'] * grouped['QTY']
    
    final_columns = [
        column_config['mark'], 'CTN NO', 'DESCRIPTION',
        'T.CTN', 'QTY', 'UNITS', 'T.QTY', 'WT', 'CONSOLIDATED_DESC'
    ]
    print(final_columns)
    print(grouped)
    final_dataset = grouped[final_columns]
    
    # Step 3: (Optional) You may reapply substitution on final_dataset if needed.
    # In this example, we assume the descriptions are already substituted.
    
    # Optionally, drop rows with a specific description.
    final_dataset = final_dataset[final_dataset['DESCRIPTION'] != "Noen"]
    
    # Save the final dataset to Excel.
    final_dataset.to_excel(output_file, index=False)
    
    # Step 4: Create a consolidated dataset.
    # Instead of grouping by the final substituted description, group by the consolidated description.
    consolidated = final_dataset.groupby('CONSOLIDATED_DESC').agg(
        Description=('CONSOLIDATED_DESC', 'first'),
        Quantity=('T.QTY', 'sum')
    )
    
    # Append the consolidated dataset to the Excel file using openpyxl.
    workbook = load_workbook(output_file)
    sheet = workbook[workbook.sheetnames[0]]
    
    # Apply styling to the entire sheet.
    cambria_font = Font(name='Cambria', size=11)
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = cambria_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write consolidated data starting at the specified cell.
    data = consolidated.reset_index(drop=True).values.tolist()
    for row_idx, row in enumerate(data, start=consolidated_start_row):
        for col_idx, value in enumerate(row, start=consolidated_start_col):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust column widths.
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(output_file)
    print("Data updated and saved successfully!")

column_config = {
    'mark': 'CUS. NO',
    'description': 'ITEM NAME',
    'pcs_per_ctn': 'Qty/ctn',
    'ctn_no': 'CTNR NO',
    'ctn_total': 'CTN',
    'weight_total':'G.W.',
    'units': 'Unit',
    # Include additional columns needed for substitution:
    # 'mah': 'MAH',
    # 'bis no.': 'BIS NO.',
    # 'bis model no.': 'BIS MODEL NO.'
}

class ColumnMapDialog(QDialog):

    def __init__(self, columns, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Map Columns")
        self.setMinimumWidth(500)
        
        # Main layout
        main_layout = QVBoxLayout()
        
        # Required column mappings
        required_fields = {
            'mark': 'Customer Number',
            'description': 'Item Name',
            'pcs_per_ctn': 'Quantity per Carton',
            'ctn_no': 'Carton Number',
            'ctn_total': 'Total Cartons',
            'weight_total': 'Gross Weight',
            'units': 'Units'
        }
        
        self.mappings = {}
        grid = QGridLayout()
        
        row = 0
        for field, display_name in required_fields.items():
            label = QLabel(display_name)
            combo = QComboBox()
            combo.addItems([''] + columns)
            combo.setObjectName(field)
            grid.addWidget(label, row, 0)
            grid.addWidget(combo, row, 1)
            self.mappings[field] = combo
            row += 1
        
        # Add the grid to the main layout
        main_layout.addLayout(grid)
        
        # Additional Columns for placeholders
        self.extra_columns_layout = QVBoxLayout()
        self.extra_columns_layout.setContentsMargins(0, 0, 0, 0)  # Remove margins for proper alignment
        
        label = QLabel("Additional Columns for Placeholders:")
        self.extra_columns_layout.addWidget(label)
        
        # Add existing extra columns if any
        self.extra_columns = []
        self.add_extra_column()
        
        # Add the extra_columns_layout to the main layout
        main_layout.addLayout(self.extra_columns_layout)
        
        # Add the "Add Column" button
        self.add_column_btn = QPushButton("Add Column")
        self.add_column_btn.clicked.connect(self.add_extra_column)
        main_layout.addWidget(self.add_column_btn)
        
        # Buttons
        buttons_layout = QHBoxLayout()
        self.ok_button = QPushButton("OK")
        self.ok_button.clicked.connect(self.accept)
        self.cancel_button = QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)
        buttons_layout.addWidget(self.ok_button)
        buttons_layout.addWidget(self.cancel_button)
        main_layout.addLayout(buttons_layout)
        
        self.setLayout(main_layout)
    
    def add_extra_column(self):
        # Create widgets for the new column
        name_edit = QLineEdit()
        name_edit.setPlaceholderText("Column Name")
        combo = QComboBox()
        
        # Get all items from the first combo box (mark)
        items = [self.mappings['mark'].itemText(i) for i in range(self.mappings['mark'].count())]
        combo.addItems(items)
        
        # Create a horizontal layout for the new column
        column_layout = QHBoxLayout()
        column_layout.addWidget(name_edit)
        column_layout.addWidget(combo)
        
        # Add the new column to the extra_columns_layout
        self.extra_columns_layout.addLayout(column_layout)
        
        # Keep track of the new column's widgets
        self.extra_columns.append((name_edit, combo))
    
    def get_mappings(self):
        result = {field: combo.currentText() for field, combo in self.mappings.items()}
        extra = {edit.text(): combo.currentText() 
                for edit, combo in self.extra_columns 
                if edit.text() and combo.currentText()}
        return result, extra

class DescriptionMappingDialog(QDialog):
    def __init__(self, missing_descriptions, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Add Missing Descriptions")
        self.setMinimumWidth(600)
        
        layout = QVBoxLayout()
        
        scroll = QScrollArea()
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout()
        
        self.mapping_inputs = {}
        for desc in missing_descriptions:
            row = QHBoxLayout()
            label = QLabel(desc)
            label.setMinimumWidth(200)
            input_field = QLineEdit()
            input_field.setMinimumWidth(300)
            row.addWidget(label)
            row.addWidget(input_field)
            self.mapping_inputs[desc] = input_field
            scroll_layout.addLayout(row)
        
        scroll_widget.setLayout(scroll_layout)
        scroll.setWidget(scroll_widget)
        scroll.setWidgetResizable(True)
        layout.addWidget(scroll)
        
        buttons = QHBoxLayout()
        ok_button = QPushButton("OK")
        ok_button.clicked.connect(self.accept)
        cancel_button = QPushButton("Cancel")
        cancel_button.clicked.connect(self.reject)
        buttons.addWidget(ok_button)
        buttons.addWidget(cancel_button)
        layout.addLayout(buttons)
        
        self.setLayout(layout)
    
    def get_mappings(self):
        return {desc: input_field.text() 
                for desc, input_field in self.mapping_inputs.items()
                if input_field.text()}

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel Processor")
        self.setMinimumSize(800, 600)
        self.setWindowIcon(QIcon("icon.png"))
        # Initialize variables
        self.df = None
        self.column_config = None
        self.json_mapping = None
        self.processed_df = None
        self.json_file_path = None
        self.selected_sheet = None  # Store the selected sheet name
        
        # Setup UI
        self.setup_ui()
        
    def setup_ui(self):
        # Create central widget and main layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)

        
        # File selection area
        file_area = QHBoxLayout()
        self.file_label = QLabel("No file selected")
        select_file_btn = QPushButton("Select Excel File")
        select_file_btn.clicked.connect(self.select_file)
        file_area.addWidget(self.file_label)
        file_area.addWidget(select_file_btn)
        layout.addLayout(file_area)
        
        # Sheet selection dropdown
        sheet_area = QHBoxLayout()
        self.sheet_label = QLabel("Select Sheet:")
        self.sheet_combo = QComboBox()
        sheet_area.addWidget(self.sheet_label)
        sheet_area.addWidget(self.sheet_combo)
        layout.addLayout(sheet_area)
        
        # JSON mapping file area
        json_area = QHBoxLayout()
        self.json_label = QLabel("No JSON mapping file selected")
        select_json_btn = QPushButton("Select JSON Mapping")
        select_json_btn.clicked.connect(self.select_json)
        json_area.addWidget(self.json_label)
        json_area.addWidget(select_json_btn)
        layout.addLayout(json_area)
        
        # Process buttons
        self.map_columns_btn = QPushButton("Map Columns")
        self.map_columns_btn.clicked.connect(self.map_columns)
        self.map_columns_btn.setEnabled(False)
        layout.addWidget(self.map_columns_btn)
        
        self.check_descriptions_btn = QPushButton("Check Descriptions")
        self.check_descriptions_btn.clicked.connect(self.check_descriptions)
        self.check_descriptions_btn.setEnabled(False)
        layout.addWidget(self.check_descriptions_btn)
        
        self.process_btn = QPushButton("Process Excel")
        self.process_btn.clicked.connect(self.process_excel)
        self.process_btn.setEnabled(False)
        layout.addWidget(self.process_btn)
        
        self.save_btn = QPushButton("Save Processed File")
        self.save_btn.clicked.connect(self.save_file)
        self.save_btn.setEnabled(False)
        layout.addWidget(self.save_btn)
        
        # Status area
        self.status_label = QLabel("")
        layout.addWidget(self.status_label)
        
        # Apply styling
        self.apply_styling()
    
    def apply_styling(self):
        # Set the application style
        self.setStyleSheet(""" 
            QMainWindow {
                background-color: #f0f0f0;
            }
            QPushButton, QComboBox {
                background-color: #2196F3;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
            QPushButton:disabled {
                background-color: #BDBDBD;
            }
            QLabel {
                color: #333333;
                font-size: 14px;
            }
        """)
    
    def select_file(self):
        file_name, _ = QFileDialog.getOpenFileName(
            self,
            "Select Excel File",
            "",
            "Excel Files (*.xlsx *.xls)"
        )
        if file_name:
            try:
                self.df = pd.read_excel(file_name, sheet_name=None)  # Read all sheets
                self.file_label.setText(f"Selected: {file_name}")
                self.sheet_combo.clear()
                self.sheet_combo.addItems(list(self.df.keys()))
                self.selected_sheet = self.sheet_combo.currentText()
                self.sheet_combo.currentTextChanged.connect(self.update_selected_sheet)
                self.map_columns_btn.setEnabled(True)
                self.status_label.setText("Excel file loaded successfully!")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error loading Excel file: {str(e)}")
    
    def update_selected_sheet(self, sheet_name):
        self.selected_sheet = sheet_name
    
    def select_json(self):
        file_name, _ = QFileDialog.getOpenFileName(
            self,
            "Select JSON Mapping File",
            "",
            "JSON Files (*.json)"
        )
        if file_name:
            self.json_file_path = file_name
            try:
                with open(file_name, 'r') as f:
                    self.json_mapping = json.load(f)
                self.json_label.setText(f"Selected: {file_name}")
                self.check_descriptions_btn.setEnabled(True)
                self.status_label.setText("JSON mapping loaded successfully!")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error loading JSON file: {str(e)}")
    
    def map_columns(self):
        if self.df is not None:
            dialog = ColumnMapDialog(list(self.df[self.selected_sheet].columns))
            if dialog.exec():
                self.column_config, extra_columns = dialog.get_mappings()
                self.column_config.update(extra_columns)
                self.status_label.setText("Columns mapped successfully!")
                self.check_descriptions_btn.setEnabled(True)
    
    def check_descriptions(self):
        if self.df is None or self.column_config is None or self.json_mapping is None:
            return
        
        # Get unique descriptions from the selected sheet
        descriptions = set(self.df[self.selected_sheet][self.column_config['description']].unique())
        
        # Get existing mappings from JSON
        existing_mappings = {item['rough'] for item in self.json_mapping}
        print(descriptions)
        print(existing_mappings)
        # Find missing descriptions
        missing_descriptions = [desc for desc in descriptions if desc not in existing_mappings and not pd.isna(desc) and desc != "" and desc.strip() != ""]
        
        if missing_descriptions:
            dialog = DescriptionMappingDialog(missing_descriptions)
            if dialog.exec():
                new_mappings = dialog.get_mappings()
                
                # Add new mappings to JSON
                for rough, formatted in new_mappings.items():
                    self.json_mapping.append({
                        "rough": rough,
                        "formatted": formatted
                    })
                
                # Save updated JSON
                with open(self.json_file_path, 'w') as f:
                    json.dump(self.json_mapping, f, indent=2)
                with open(self.json_file_path, 'r') as f:
                    self.json_mapping = json.load(f)
                self.status_label.setText("Descriptions updated successfully!")
                self.process_btn.setEnabled(True)
        else:
            self.status_label.setText("All descriptions found in mapping!")
            self.process_btn.setEnabled(True)
    
    def process_excel(self):
        if all([self.df is not None, self.column_config is not None, self.json_mapping is not None]):
            try:
                # Call your existing process_data function here
                process_data(
                    df=self.df[self.selected_sheet],  # Use the selected sheet
                    sheet_name=self.selected_sheet,
                    output_file='temp_output.xlsx',
                    column_config=self.column_config,
                    description_mapping=self.json_mapping
                )
                self.status_label.setText("Excel processed successfully!")
                self.save_btn.setEnabled(True)
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error processing Excel: {str(e)}")
    
    def save_file(self):
        file_name, _ = QFileDialog.getSaveFileName(
            self,
            "Save Processed File",
            "",
            "Excel Files (*.xlsx)"
        )
        if file_name:
            try:
                # Copy the processed file to the selected location
                import shutil
                shutil.copy2('temp_output.xlsx', file_name)
                self.status_label.setText(f"File saved successfully to: {file_name}")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error saving file: {str(e)}")

def main():
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())

if __name__ == '__main__':
    main()