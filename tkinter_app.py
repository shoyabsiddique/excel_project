import pandas as pd
import json
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

class GenericExcelProcessor:
    def __init__(self, root):
        self.root = root
        self.root.title("Universal Excel Processor")
        
        self.file_path = None
        self.df = None
        self.mappings = {}
        self.config = {
            'group_columns': [],
            'block_column': None,
            'aggregations': {},
            'output_columns': [],
            'replace_column': None
        }
        
        self.create_ui()
        self.load_config()
    
    def create_ui(self):
        main_frame = ttk.Frame(self.root, padding=20)
        main_frame.pack(fill='both', expand=True)
        
        ttk.Button(main_frame, text="Load Excel File", command=self.load_file).pack(pady=10)
        
        self.setup_frame = ttk.LabelFrame(main_frame, text="Processing Configuration")
        self.setup_frame.pack(fill='x', pady=10)
        
        ttk.Button(main_frame, text="Run Processing", command=self.run_processing).pack(pady=10)
    
    def load_config(self):
        try:
            with open('processor_config.json') as f:
                self.config = json.load(f)
        except FileNotFoundError:
            pass
    
    def save_config(self):
        with open('processor_config.json', 'w') as f:
            json.dump(self.config, f)
    
    def load_file(self):
        self.file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if self.file_path:
            try:
                self.df = pd.read_excel(self.file_path)
                self.show_config_interface()
            except Exception as e:
                messagebox.showerror("Error", f"Error loading file: {str(e)}")
    
    def show_config_interface(self):
        for widget in self.setup_frame.winfo_children():
            widget.destroy()
            
        cols = self.df.columns.tolist()
        
        # Grouping Configuration
        ttk.Label(self.setup_frame, text="Grouping Columns:").grid(row=0, column=0, sticky='w')
        self.group_vars = [tk.StringVar() for _ in range(3)]
        for i in range(3):
            ttk.Combobox(self.setup_frame, textvariable=self.group_vars[i], values=cols).grid(row=0, column=i+1)
        
        # Block Column
        ttk.Label(self.setup_frame, text="Block Detection Column:").grid(row=1, column=0, sticky='w')
        self.block_var = tk.StringVar()
        ttk.Combobox(self.setup_frame, textvariable=self.block_var, values=cols).grid(row=1, column=1)
        
        # Aggregation Setup
        ttk.Label(self.setup_frame, text="Column Aggregations:").grid(row=2, column=0, sticky='w')
        self.agg_vars = {}
        for i, col in enumerate(cols):
            ttk.Label(self.setup_frame, text=col).grid(row=3+i, column=0)
            self.agg_vars[col] = tk.StringVar()
            ttk.Combobox(self.setup_frame, textvariable=self.agg_vars[col], 
                        values=['first', 'last', 'sum', 'count', 'mean']).grid(row=3+i, column=1)
        
        # Replacement Column
        ttk.Label(self.setup_frame, text="Description Replacement Column:").grid(row=2, column=2, sticky='w')
        self.replace_var = tk.StringVar()
        ttk.Combobox(self.setup_frame, textvariable=self.replace_var, values=cols).grid(row=2, column=3)
    
    def run_processing(self):
        try:
            self.save_configuration()
            processed_df = self.process_data()
            self.handle_replacements(processed_df)
            self.save_output(processed_df)
        except Exception as e:
            messagebox.showerror("Processing Error", str(e))
    
    def save_configuration(self):
        self.config['group_columns'] = [v.get() for v in self.group_vars if v.get()]
        self.config['block_column'] = self.block_var.get()
        self.config['aggregations'] = {col: var.get() for col, var in self.agg_vars.items()}
        self.config['replace_column'] = self.replace_var.get()
        self.save_config()
    
    def process_data(self):
        df = self.df
        
        # Create blocks
        if self.config['block_column']:
            df['Block'] = (df[self.config['block_column']] != df[self.config['block_column']].shift()).cumsum()
            group_columns = self.config['group_columns'] + ['Block']
        else:
            group_columns = self.config['group_columns']
        
        # Perform aggregations
        agg_dict = {}
        for col, agg_func in self.config['aggregations'].items():
            if agg_func:
                agg_dict[col] = (col, agg_func)
        
        grouped = df.groupby(group_columns).agg(**agg_dict).reset_index()
        
        # Generate CTN-like fields
        if 'CTN First-Last' in agg_dict.values():
            grouped['CTN Range'] = grouped.apply(
                lambda row: f"{row['First_CTN']}-{row['Last_CTN']}" 
                if row['First_CTN'] != row['Last_CTN'] 
                else str(row['First_CTN']), axis=1
            )
        
        return grouped
    
    def handle_replacements(self, df):
        if not self.config['replace_column']:
            return
            
        try:
            with open('replacements.json') as f:
                replacements = json.load(f)
        except FileNotFoundError:
            replacements = {}
            
        new_entries = {}
        for value in df[self.config['replace_column']].unique():
            if value not in replacements:
                new_name = self.get_replacement_name(value)
                if new_name:
                    new_entries[value] = new_name
                    
        replacements.update(new_entries)
        df[self.config['replace_column']] = df[self.config['replace_column']].replace(replacements)
        
        with open('replacements.json', 'w') as f:
            json.dump(replacements, f)
    
    def get_replacement_name(self, original):
        popup = tk.Toplevel()
        popup.title("New Entry Required")
        
        ttk.Label(popup, text=f"Replacement for '{original}':").pack(pady=5)
        entry = ttk.Entry(popup)
        entry.pack(pady=5)
        
        result = [None]
        def save():
            result[0] = entry.get()
            popup.destroy()
        
        ttk.Button(popup, text="Save", command=save).pack(pady=5)
        popup.transient(self.root)
        popup.grab_set()
        self.root.wait_window(popup)
        return result[0]
    
    def save_output(self, df):
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if save_path:
            df.to_excel(save_path, index=False)
            self.apply_formatting(save_path)
            messagebox.showinfo("Success", "File processed successfully!")
    
    def apply_formatting(self, path):
        wb = load_workbook(path)
        ws = wb.active
        
        font = Font(name='Calibri', size=11)
        align = Alignment(horizontal="center", vertical="center")
        
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font
                cell.alignment = align
        
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2
        
        wb.save(path)

if __name__ == "__main__":
    root = tk.Tk()
    app = GenericExcelProcessor(root)
    root.mainloop()