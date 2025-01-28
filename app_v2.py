import os
import pandas as pd
import json
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# Load names.json
def load_name_mapping(file_path):
    # Check if the file exists
    if not os.path.exists(file_path):
        # If the file doesn't exist, create it with an empty list
        with open(file_path, 'w') as file:
            json.dump([], file, indent=4)
    # Load the JSON file
    with open(file_path, 'r') as file:
        return json.load(file)

# Save the name mapping back to the file
def save_name_mapping(file_path, name_mapping):
    with open(file_path, 'w') as file:
        json.dump(name_mapping, file, indent=4)

# Function to process the dataset
def process_dataset(dataframe, name_mapping, column_mapping):
    dataframe.dropna(inplace=True)
    description_mapping = {item['rough']: item['formatted'] for item in name_mapping}
    dataframe[column_mapping['DESCRIPTION']] = dataframe[column_mapping['DESCRIPTION']].replace(description_mapping)
    dataframe['Block'] = (dataframe[column_mapping['PCS/CTN']] != dataframe[column_mapping['PCS/CTN']].shift()).cumsum()
    grouped = dataframe.groupby([column_mapping['MARK'], 'Block']).agg(
        First_CTN=(column_mapping['CTN NO'], 'first'),
        Last_CTN=(column_mapping['CTN NO'], 'last'),
        T_CTN=(column_mapping['CTN/TOTAL'], 'sum'),
        WT=(column_mapping['WEIGHT/TOTAL'], 'first'),
        UNIT=(column_mapping['UNITS'], 'first'),
        QTY=(column_mapping['PCS/CTN'], 'first'),
        T_QTY=(column_mapping['PCS/CTN'], 'sum'),
        DESCRIPTION=(column_mapping['DESCRIPTION'], 'first')
    ).reset_index()

    grouped['CTN NO'] = grouped.apply(
        lambda row: (
            f"{row['First_CTN'] if row['First_CTN'] is not None else 1} - "
            f"{row['Last_CTN'] if row['Last_CTN'] is not None else int(row['T_CTN'])}"
            if row['T_CTN'] > 1
            else f"{row['First_CTN'] if row['First_CTN'] is not None else 1}"
        ),
        axis=1
    )
    grouped['T.QTY'] = grouped['T_QTY']
    grouped['T.CTN'] = grouped['T_CTN']
    final_columns = ['MARK', 'CTN NO', 'DESCRIPTION', 'T.CTN', 'QTY', 'UNIT', 'T.QTY', 'WT']
    final_dataset = grouped[final_columns]
    return final_dataset

# Main Streamlit app
st.title("Excel Processing App")

# Step 1: Upload Excel file
uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])

if uploaded_file:
    dataframe = pd.read_excel(uploaded_file)
    st.write("Preview of uploaded data:")
    st.dataframe(dataframe.head())

    # Step 2: Column mapping
    column_mapping = {}
    column_mapping['MARK'] = st.selectbox("Select the MARK column", dataframe.columns.tolist())
    column_mapping['CTN NO'] = st.selectbox("Select the CTN NO column", dataframe.columns.tolist())
    column_mapping['DESCRIPTION'] = st.selectbox("Select the DESCRIPTION column", dataframe.columns.tolist())
    column_mapping['CTN/TOTAL'] = st.selectbox("Select the CTN/TOTAL column", dataframe.columns.tolist())
    column_mapping['WEIGHT/TOTAL'] = st.selectbox("Select the WEIGHT/TOTAL column", dataframe.columns.tolist())
    column_mapping['UNITS'] = st.selectbox("Select the UNITS column", dataframe.columns.tolist())
    column_mapping['PCS/CTN'] = st.selectbox("Select the PCS/CTN column", dataframe.columns.tolist())

    if all(column_mapping.values()):
        st.write("Preview of selected columns:")
        st.dataframe(dataframe[column_mapping.values()].head())

        # Step 3: Process the dataset
        name_mapping = load_name_mapping("names.json")
        processed_dataset = process_dataset(dataframe, name_mapping, column_mapping)

        # Step 4: Check for missing names
        missing_names = processed_dataset[~processed_dataset['DESCRIPTION'].isin(
            [item['formatted'] for item in name_mapping]
        )]['DESCRIPTION'].unique()

        if missing_names.size > 0:
            st.warning("Some descriptions are not mapped:")
            for name in missing_names:
                new_name = st.text_input(f"Enter formatted name for '{name}':")
                if new_name:
                    # Append the new name to the name mapping
                    name_mapping.append({"rough": name, "formatted": new_name})
                    save_name_mapping("names.json", name_mapping)
                    st.success(f"'{name}' mapped to '{new_name}' and saved.")

        # Step 5: Display processed data
        st.write("Preview of processed data:")
        st.dataframe(processed_dataset)

        # Step 6: Save processed data
        save_path = st.text_input("Enter a filename to save the processed file (e.g., processed.xlsx):")
        if st.button("Save Processed Data"):
            processed_dataset.to_excel(save_path, index=False)
            st.success(f"Processed file saved as {save_path}!")

            consolidated = processed_dataset.groupby(['DESCRIPTION'])
            consolidated_dataset = consolidated.agg(
                Description=('DESCRIPTION', 'first'),
                Quantity=('T.QTY', 'sum'),
            )

            # Write the consolidated data back to the Excel file
            # Append the consolidated dataset to the Excel file as before
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Alignment

            # Load the existing Excel file
            file_path = save_path  # Replace with your file path
            workbook = load_workbook(file_path)

            # Load the first sheet (or specify the sheet name)
            sheet_name = workbook.sheetnames[0]
            sheet = workbook[sheet_name]

            cambria_font = Font(name='Cambria', size=11)
            for row in sheet.iter_rows():
                for cell in row:
                    # Set the font
                    cell.font = cambria_font
                    cell.border = None
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Define the starting cell for consolidated data
            start_row = 6
            start_col = 10  # Column 'J' corresponds to the 10th column
            data = consolidated_dataset.values.tolist()
            headers = consolidated_dataset.columns.tolist()

            # Write data
            for row_idx, row in enumerate(data, start=start_row):
                for col_idx, value in enumerate(row, start=start_col):
                    sheet.cell(row=row_idx, column=col_idx, value=value)

            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter  # Get column letter (e.g., A, B, C)
                for cell in column:
                    if cell.value:  # Check if cell is not empty
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = max_length + 2  # Add extra space for padding
                sheet.column_dimensions[column_letter].width = adjusted_width

            # Save the workbook
            workbook.save(save_path)

            print("Data updated and saved successfully!")
