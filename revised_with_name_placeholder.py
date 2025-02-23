import pandas as pd
import numpy as np
import json
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

def process_data(
    input_file,
    sheet_name,
    output_file,
    column_config,
    json_mapping_file,
    consolidated_start_row=6,
    consolidated_start_col=10,
    drop_columns=None
):
    """
    Process the Excel data dynamically based on the user-specified column configuration.
    Instead of dropping rows with missing values, this function fills them with default values.
    
    It also handles placeholder substitution in the description column. When the JSON mapping 
    contains placeholders (like %BIS NO.% or %MAH%), the substitution function returns:
      - the substituted description (for display)
      - the original rough description (for consolidation) 
    so that rows coming from the same rough mapping are consolidated together.
    
    Parameters:
    - input_file: Path to the input Excel file.
    - sheet_name: Sheet name in the Excel file to read.
    - output_file: Path where the processed Excel file will be saved.
    - column_config: Dictionary mapping roles to column names. For example:
        {
            'mark': 'MARK',
            'description': 'DESCRIPTION',
            'pcs_per_ctn': 'PCS/CTN',
            'ctn_no': 'CTN NO',
            'ctn_total': 'CTN/TOTAL',
            'weight_total': 'WEIGHT/TOTAL',
            'units': 'UNITS'
        }
      **Important:** Include any additional columns (e.g. 'MAH', 'BIS NO.', 'BIS MODEL NO.') needed for placeholder substitution.
    - json_mapping_file: Path to the JSON file containing description mappings.
    - consolidated_start_row: Row number where consolidated data starts (default 6).
    - consolidated_start_col: Column index where consolidated data starts (default 10 for 'J').
    - drop_columns: List of columns to drop from the input DataFrame.
    """
    
    # Step 1: Read the Excel file and drop unwanted columns.
    df = pd.read_excel(input_file, sheet_name=sheet_name)
    if drop_columns:
        df = df.drop(columns=drop_columns)
    
    # Fill missing values for the mark column.
    df[column_config['mark']] = df[column_config['mark']].fillna("Unknown")
    df[column_config['description']] = df[column_config['description']].fillna("None")
    df = df[df[column_config['description']] != "None"]
    # For any extra columns needed for placeholder substitution (like MAH),
    # fill missing values if desired.
    if 'mah' in column_config:
        df[column_config['mah']] = df[column_config['mah']].fillna("0")
    
    # Load the JSON mapping file.
    with open(json_mapping_file, 'r') as file:
        name_mapping = json.load(file)
    
    # Build a dictionary mapping rough description to the formatted string.
    # (No normalization is applied here, so ensure the keys match exactly.)
    description_mapping = {item['rough']: item['formatted'] for item in name_mapping}
    
    # Define a substitution function.
    def substitute_placeholders(row, mapping, col_config):
        """
        Returns a tuple: (final_description, consolidated_description)
        - final_description: The result of performing placeholder substitution.
        - consolidated_description: If any placeholders were used, then return the original rough description,
          otherwise return the substituted string.
        """
        raw = row[col_config['description']]
        if raw in mapping:
            formatted = mapping[raw]
            # Use a regex that captures anything between % signs.
            placeholders = re.findall(r'%([^%]+)%', formatted)
            new_desc = formatted
            for ph in placeholders:
                # Replace only if the column exists in the row.
                if ph in row:
                    new_desc = new_desc.replace(f'%{ph}%', str(row[ph]))
            # If there were placeholders, then for consolidation use the raw (rough) description.
            if placeholders:
                return new_desc, raw
            else:
                return new_desc, new_desc
        return raw, raw

    # Before grouping, apply substitution so that we create two new columns:
    # - one for the final (display) description
    # - one for consolidation (rough) description.
    results = df.apply(lambda row: pd.Series(substitute_placeholders(row, description_mapping, column_config)), axis=1)
    # Assign the two new columns.
    df[column_config['description']] = results[0]
    df['consolidated_desc'] = results[1]
    
    # Create a 'Block' identifier based on changes in the 'pcs_per_ctn' column.
    df['Block'] = (df[column_config['pcs_per_ctn']] != df[column_config['pcs_per_ctn']].shift()).cumsum()
    
    # Step 2: Group by the 'mark' and 'Block', then aggregate.
    # Include both the final description and the consolidated description in the aggregation.
    grouped = df.groupby([column_config['mark'], 'Block']).agg(
        First_CTN=(column_config['ctn_no'], 'first'),
        Last_CTN=(column_config['ctn_no'], 'last'),
        T_CTN=(column_config['ctn_total'], 'sum'),
        WT=(column_config['weight_total'], 'first'),
        UNITS=(column_config['units'], 'first'),
        QTY=(column_config['pcs_per_ctn'], 'first'),
        T_QTY=(column_config['pcs_per_ctn'], 'sum'),
        DESCRIPTION=(column_config['description'], 'first'),
        CONSOLIDATED_DESC=('consolidated_desc', 'first')
    ).reset_index()
    
    # Generate the CTN NO field dynamically.
    grouped['CTN NO'] = grouped.apply(
        lambda row: (
            f"{row['First_CTN'] if pd.notnull(row['First_CTN']) else 1} - "
            f"{row['Last_CTN'] if pd.notnull(row['Last_CTN']) else int(row['T_CTN'])}"
            if row['T_CTN'] > 1
            else f"{row['First_CTN'] if pd.notnull(row['First_CTN']) else 1}"
        ),
        axis=1
    )
    
    # Rename columns for the final dataset.
    grouped['T.QTY'] = grouped['T_QTY']
    grouped['T.CTN'] = grouped['T_CTN']
    
    final_columns = [
        column_config['mark'], 'CTN NO', 'DESCRIPTION',
        'T.CTN', 'QTY', 'UNITS', 'T.QTY', 'WT', 'CONSOLIDATED_DESC'
    ]
    print(final_columns)
    print(grouped)
    final_dataset = grouped[final_columns]
    
    # Step 3: (Optional) You may reapply substitution on final_dataset if needed.
    # In this example, we assume the descriptions are already substituted.
    
    # Optionally, drop rows with a specific description.
    final_dataset = final_dataset[final_dataset['DESCRIPTION'] != "Noen"]
    
    # Save the final dataset to Excel.
    final_dataset.to_excel(output_file, index=False)
    
    # Step 4: Create a consolidated dataset.
    # Instead of grouping by the final substituted description, group by the consolidated description.
    consolidated = final_dataset.groupby('CONSOLIDATED_DESC').agg(
        Description=('CONSOLIDATED_DESC', 'first'),
        Quantity=('T.QTY', 'sum')
    )
    
    # Append the consolidated dataset to the Excel file using openpyxl.
    workbook = load_workbook(output_file)
    sheet = workbook[workbook.sheetnames[0]]
    
    # Apply styling to the entire sheet.
    cambria_font = Font(name='Cambria', size=11)
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = cambria_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write consolidated data starting at the specified cell.
    data = consolidated.reset_index(drop=True).values.tolist()
    for row_idx, row in enumerate(data, start=consolidated_start_row):
        for col_idx, value in enumerate(row, start=consolidated_start_col):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust column widths.
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(output_file)
    print("Data updated and saved successfully!")

# Example usage:

# Define the column configuration mapping.
# Ensure that if you want to use placeholders like %MAH%, %BIS NO.% etc.,
# the corresponding columns must be kept.
column_config = {
    'mark': 'CUS. NO',
    'description': 'ITEM NAME',
    'pcs_per_ctn': 'Qty/ctn',
    'ctn_no': 'CTNR NO',
    'ctn_total': 'CTN',
    'weight_total':'G.W.',
    'units': 'Unit',
    # Include additional columns needed for substitution:
    # 'mah': 'MAH',
    # 'bis no.': 'BIS NO.',
    # 'bis model no.': 'BIS MODEL NO.'
}
# column_config = {
#     'mark': 'MARK',
#     'description': 'DESCRIPTION',
#     'pcs_per_ctn': 'PCS/CTN',
#     'ctn_no': 'CTN NO',
#     'ctn_total': 'PCS/TOTAL',
#     'weight_total':'WEIGHT/TOTAL',
#     'units': 'UNITS',
#     # Include additional columns needed for substitution:
#     'mah': 'MAH',
#     'bis no.': 'BIS NO.',
#     'bis model no.': 'BIS MODEL NO.'
# }

# Adjust drop_columns as needed. In this example, we ensure that the 'MAH' column is not dropped.
# drop_columns = ['BIS NO.', 'BIS MODEL NO.', 'MADE IN', 'LOGO']  # 'MAH' is kept for substitution

process_data(
    input_file='FIN-RAN-H11-NS.xlsx',
    sheet_name='H11-NS',
    output_file='processed_revised_new_generic.xlsx',
    column_config=column_config,
    json_mapping_file='names1.json',
    consolidated_start_row=6,
    consolidated_start_col=10,
    drop_columns=None
)
