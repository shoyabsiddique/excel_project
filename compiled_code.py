import pandas as pd
import numpy as np
dataframe = pd.read_excel('test.xlsx', sheet_name='Sheet1')
dataframe.head()
dataframe = dataframe.drop(columns=['BIS NO.', 'BIS MODEL NO.', 'MAH', 'MADE IN', 'LOGO'])
dataframe.dropna(inplace=True, subset=['MARK'])
dataframe.head()
grouped = dataframe.groupby(['MARK', 'DESCRIPTION', 'PCS/CTN'])
print(grouped.head(10))
dataframe['Block'] = (dataframe['PCS/CTN'] != dataframe['PCS/CTN'].shift()).cumsum()

# Step 2: Group by MARK, Block, and generate summaries
grouped = dataframe.groupby(['MARK', 'Block']).agg(
    First_CTN=('CTN NO', 'first'),
    Last_CTN=('CTN NO', 'last'),
    T_CTN=('CTN/TOTAL', 'sum'),
    WT=('WEIGHT/TOTAL', 'first'),
    UNIT=('UNITS', 'first'),
    QTY=('PCS/CTN', 'first'),
    T_QTY=('PCS/CTN', 'sum'),
    DESCRIPTION=('DESCRIPTION', 'first')
).reset_index()

# Step 3: Generate the CTN NO field
grouped['CTN NO'] = grouped.apply(
    lambda row: (
        f"{row['First_CTN'] if row['First_CTN'] is not None else 1} - "
        f"{row['Last_CTN'] if row['Last_CTN'] is not None else int(row['T_CTN'])}"
        if row['T_CTN'] > 1
        else f"{row['First_CTN'] if row['First_CTN'] is not None else 1}"
    ),
    axis=1
)
grouped['T.QTY'] = grouped['T_QTY']
grouped['T.CTN'] = grouped['T_CTN']
# Step 4: Reorder and clean up
final_columns = ['MARK', 'CTN NO', 'DESCRIPTION', 'T.CTN', 'QTY', 'UNIT', 'T.QTY', 'WT']
# grouped['DESCRIPTION'] = 'BACK COVER'  # Assuming the same description
final_dataset = grouped[final_columns]
import json

# Load the JSON file
with open('names.json', 'r') as file:
    name_mapping = json.load(file)

# Create a mapping dictionary from the JSON data
description_mapping = {item['rough']: item['formatted'] for item in name_mapping}

# Replace the DESCRIPTION column values using the mapping
final_dataset['DESCRIPTION'] = final_dataset['DESCRIPTION'].replace(description_mapping)

# Display the updated dataset
print(final_dataset)

# Save the updated dataset to the Excel file
final_dataset.to_excel('processed.xlsx', index=False)

# Consolidated dataset also needs updated descriptions
consolidated = final_dataset.groupby(['DESCRIPTION'])
consolidated_dataset = consolidated.agg(
    Description=('DESCRIPTION', 'first'),
    Quantity=('T.QTY', 'sum'),
)

# Write the consolidated data back to the Excel file
# Append the consolidated dataset to the Excel file as before
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# Load the existing Excel file
file_path = 'processed.xlsx'  # Replace with your file path
workbook = load_workbook(file_path)

# Load the first sheet (or specify the sheet name)
sheet_name = workbook.sheetnames[0]
sheet = workbook[sheet_name]

cambria_font = Font(name='Cambria', size=11)
for row in sheet.iter_rows():
    for cell in row:
        # Set the font
        cell.font = cambria_font
        cell.border = None
        cell.alignment = Alignment(horizontal="center", vertical="center")

# Define the starting cell for consolidated data
start_row = 6
start_col = 10  # Column 'J' corresponds to the 10th column
data = consolidated_dataset.values.tolist()
print(data)
headers = consolidated_dataset.columns.tolist()

# Write data
for row_idx, row in enumerate(data, start=start_row):
    for col_idx, value in enumerate(row, start=start_col):
        sheet.cell(row=row_idx, column=col_idx, value=value)

# Auto-adjust column widths
for column in sheet.columns:
    max_length = 0
    column_letter = column[0].column_letter  # Get column letter (e.g., A, B, C)
    for cell in column:
        if cell.value:  # Check if cell is not empty
            max_length = max(max_length, len(str(cell.value)))
    adjusted_width = max_length + 2  # Add extra space for padding
    sheet.column_dimensions[column_letter].width = adjusted_width

# Save the workbook
workbook.save(file_path)

print("Data updated and saved successfully!")

# # Display the final dataset
# print(final_dataset)
# final_dataset.to_excel('processed.xlsx', index=False)

# consolidated = final_dataset.groupby(['DESCRIPTION'])
# consolidated_dataset = consolidated.agg(
#     Description=('DESCRIPTION', 'first'),
#     Quantity=('T.QTY', 'sum'),
# )
# consolidated_dataset

# from openpyxl import load_workbook
# from openpyxl.styles import Font, Alignment

# # Load the existing Excel file
# file_path = 'processed.xlsx'  # Replace with your file path
# workbook = load_workbook(file_path)


# # Load the first sheet (or specify the sheet name)
# sheet_name = workbook.sheetnames[0]
# sheet = workbook[sheet_name]
# cambria_font = Font(name='Cambria', size=11)
# for row in sheet.iter_rows():
#     for cell in row:
#         # Set the font
#         cell.font = cambria_font
#         # Remove borders by setting none (optional, usually default)
#         cell.border = None
#         # Align content centrally (optional, adjust as needed)
#         cell.alignment = Alignment(horizontal="center", vertical="center")
# # Define the starting cell
# start_row = 6
# start_col = 10  # Column 'J' corresponds to the 10th column
# data = consolidated_dataset.values.tolist()
# print(data)
# headers = consolidated_dataset.columns.tolist()

# # Write data
# for row_idx, row in enumerate(data, start=start_row):
#     for col_idx, value in enumerate(row, start=start_col):
#         sheet.cell(row=row_idx, column=col_idx, value=value)
# # Auto-adjust column widths
# for column in sheet.columns:
#     max_length = 0
#     column_letter = column[0].column_letter  # Get column letter (e.g., A, B, C)
#     for cell in column:
#         if cell.value:  # Check if cell is not empty
#             max_length = max(max_length, len(str(cell.value)))
#     adjusted_width = max_length + 2  # Add extra space for padding
#     sheet.column_dimensions[column_letter].width = adjusted_width
# # Save the workbook
# workbook.save(file_path)

# print("Data appended successfully!")